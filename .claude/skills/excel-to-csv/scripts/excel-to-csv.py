#!/usr/bin/env python3
"""Excel/CSV to UTF-8 CSV converter for Claude Code.

Usage:
  python excel-to-csv.py <file_or_folder> [options]

Options:
  --info              Show sheet list and preview only (no conversion)
  --sheet <name>      Convert specific sheet only
  --output <path>     Output directory (default: same as source)
  --all               Convert all sheets without confirmation
  --encoding <enc>    Force source encoding (for CSV input)
"""

import argparse
import csv
import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl>=3.1.0", file=sys.stderr)
    sys.exit(1)


def sanitize_filename(name: str) -> str:
    """Sanitize filename: keep Korean, replace spaces, remove special chars."""
    name = name.replace(" ", "_")
    name = re.sub(r'[<>:"/\\|?*]', "", name)
    return name


def detect_encoding(file_path: str) -> str:
    """Detect file encoding by trying common Korean encodings."""
    encodings = ["utf-8", "utf-8-sig", "euc-kr", "cp949", "utf-16", "utf-16-le"]
    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc) as f:
                f.read(4096)
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return "utf-8"


def find_header_row(ws) -> int:
    """Find the first non-empty row (0-indexed) as header."""
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if any(cell is not None and str(cell).strip() for cell in row):
            return idx
    return 0


def get_max_col(ws) -> int:
    """Find the last non-empty column."""
    max_col = 0
    for row in ws.iter_rows(values_only=False):
        for cell in reversed(row):
            if cell.value is not None and str(cell.value).strip():
                max_col = max(max_col, cell.column)
                break
    return max_col


def get_max_row(ws, header_row: int) -> int:
    """Find the last non-empty row after header."""
    max_row = header_row
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx >= header_row and any(cell is not None and str(cell).strip() for cell in row):
            max_row = idx
    return max_row


def handle_merged_cells(ws):
    """Fill merged cell regions with the top-left value."""
    for merge_range in list(ws.merged_cells.ranges):
        min_row = merge_range.min_row
        min_col = merge_range.min_col
        value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merge_range))
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                ws.cell(row=row, column=col).value = value


def sheet_info(ws, sheet_name: str) -> dict:
    """Get sheet metadata."""
    header_row = find_header_row(ws)
    max_col = get_max_col(ws)
    max_row = get_max_row(ws, header_row)
    data_rows = max_row - header_row  # excluding header

    if max_col == 0 or data_rows <= 0:
        return {"name": sheet_name, "rows": 0, "cols": 0, "headers": [], "empty": True}

    headers = []
    for cell in list(ws.iter_rows(min_row=header_row + 1, max_row=header_row + 1,
                                   min_col=1, max_col=max_col, values_only=True))[0]:
        headers.append(str(cell) if cell is not None else "")

    preview_rows = []
    row_iter = ws.iter_rows(min_row=header_row + 2, max_row=min(header_row + 6, max_row + 1),
                            min_col=1, max_col=max_col, values_only=True)
    for row in row_iter:
        preview_rows.append([str(c) if c is not None else "" for c in row])

    return {
        "name": sheet_name,
        "rows": data_rows,
        "cols": max_col,
        "headers": headers,
        "preview": preview_rows,
        "empty": False,
    }


def print_info(file_path: str):
    """Print sheet information for an Excel file."""
    wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
    print(f"File: {file_path}")
    print(f"Sheets: {len(wb.sheetnames)}")
    print("-" * 60)

    for name in wb.sheetnames:
        ws = wb[name]
        info = sheet_info(ws, name)
        if info["empty"]:
            print(f"  [{name}] (empty - skipped)")
            continue
        print(f"  [{name}] {info['rows']} rows x {info['cols']} cols")
        print(f"    Headers: {', '.join(info['headers'][:10])}")
        if len(info["headers"]) > 10:
            print(f"    ... and {len(info['headers']) - 10} more columns")
        if info.get("preview"):
            print(f"    Preview (first {len(info['preview'])} rows):")
            for row in info["preview"][:3]:
                truncated = [v[:30] + "..." if len(v) > 30 else v for v in row[:6]]
                print(f"      {' | '.join(truncated)}")
        print()
    wb.close()


def convert_sheet_to_csv(ws, output_path: str, sheet_name: str) -> dict:
    """Convert a single worksheet to CSV. Returns metadata."""
    handle_merged_cells(ws)
    header_row = find_header_row(ws)
    max_col = get_max_col(ws)
    max_row = get_max_row(ws, header_row)
    data_rows = max_row - header_row

    if max_col == 0 or data_rows <= 0:
        return {"sheet": sheet_name, "rows": 0, "skipped": True}

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(min_row=header_row + 1, max_row=max_row + 1,
                                min_col=1, max_col=max_col, values_only=True):
            writer.writerow([str(c) if c is not None else "" for c in row])

    return {"sheet": sheet_name, "rows": data_rows, "path": output_path, "skipped": False}


def convert_excel(file_path: str, output_dir: str, sheet_name: str = None, convert_all: bool = False):
    """Convert Excel file to CSV(s)."""
    wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
    stem = sanitize_filename(Path(file_path).stem)
    results = []

    sheets_to_convert = []
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"ERROR: Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}", file=sys.stderr)
            wb.close()
            sys.exit(1)
        sheets_to_convert = [sheet_name]
    else:
        sheets_to_convert = wb.sheetnames

    # Determine naming: single sheet = simple name, multi = with sheet suffix
    use_suffix = len(sheets_to_convert) > 1

    for name in sheets_to_convert:
        ws = wb[name]
        if use_suffix:
            csv_name = f"{stem}_{sanitize_filename(name)}.csv"
        else:
            csv_name = f"{stem}.csv"

        out_path = os.path.join(output_dir, csv_name)
        result = convert_sheet_to_csv(ws, out_path, name)
        results.append(result)

    wb.close()

    # Print results
    print(f"Converted: {Path(file_path).name}")
    converted_count = 0
    total_rows = 0
    for r in results:
        if r.get("skipped"):
            print(f"  [{r['sheet']}] skipped (empty)")
        else:
            print(f"  [{r['sheet']}] -> {Path(r['path']).name} ({r['rows']} rows)")
            converted_count += 1
            total_rows += r["rows"]
    print(f"Total: {converted_count} file(s), {total_rows} rows")
    return results


def convert_csv_encoding(file_path: str, output_dir: str, source_encoding: str = None):
    """Convert a CSV file from detected/specified encoding to UTF-8."""
    if source_encoding is None:
        source_encoding = detect_encoding(file_path)

    stem = sanitize_filename(Path(file_path).stem)
    out_path = os.path.join(output_dir, f"{stem}_utf8.csv")

    with open(file_path, "r", encoding=source_encoding) as infile:
        content = infile.read()

    with open(out_path, "w", encoding="utf-8", newline="") as outfile:
        outfile.write(content)

    # Count rows
    row_count = content.count("\n")
    print(f"Converted: {Path(file_path).name}")
    print(f"  Encoding: {source_encoding} -> UTF-8")
    print(f"  Output: {Path(out_path).name} ({row_count} rows)")
    return [{"sheet": "csv", "rows": row_count, "path": out_path, "skipped": False}]


def process_folder(folder_path: str, output_dir: str, convert_all: bool = False):
    """Process all Excel files in a folder."""
    folder = Path(folder_path)
    excel_files = sorted(list(folder.glob("*.xlsx")) + list(folder.glob("*.xls")) + list(folder.glob("*.XLSX")))

    if not excel_files:
        print(f"No Excel files found in {folder_path}")
        return

    print(f"Found {len(excel_files)} Excel file(s) in {folder_path}")
    print("-" * 60)

    all_results = []
    for fp in excel_files:
        results = convert_excel(str(fp), output_dir, convert_all=convert_all)
        all_results.extend(results)
        print()

    total_files = sum(1 for r in all_results if not r.get("skipped"))
    total_rows = sum(r.get("rows", 0) for r in all_results if not r.get("skipped"))
    print("=" * 60)
    print(f"Grand total: {len(excel_files)} Excel file(s) -> {total_files} CSV file(s), {total_rows} rows")


def main():
    parser = argparse.ArgumentParser(description="Excel/CSV to UTF-8 CSV converter")
    parser.add_argument("path", help="Excel file or folder path")
    parser.add_argument("--info", action="store_true", help="Show sheet info only")
    parser.add_argument("--sheet", help="Convert specific sheet only")
    parser.add_argument("--output", help="Output directory")
    parser.add_argument("--all", action="store_true", dest="convert_all", help="Convert all sheets")
    parser.add_argument("--encoding", help="Force source encoding (for CSV files)")

    args = parser.parse_args()
    target = os.path.abspath(args.path)

    if not os.path.exists(target):
        print(f"ERROR: Path not found: {target}", file=sys.stderr)
        sys.exit(1)

    # Determine output directory
    if args.output:
        output_dir = os.path.abspath(args.output)
        os.makedirs(output_dir, exist_ok=True)
    elif os.path.isfile(target):
        output_dir = os.path.dirname(target)
    else:
        output_dir = target

    # Folder mode
    if os.path.isdir(target):
        if args.info:
            for fp in sorted(Path(target).glob("*.xlsx")) + sorted(Path(target).glob("*.XLSX")):
                print_info(str(fp))
                print()
        else:
            process_folder(target, output_dir, args.convert_all)
        return

    # Single file mode
    ext = Path(target).suffix.lower()

    if ext in (".xlsx", ".xls"):
        if args.info:
            print_info(target)
        else:
            convert_excel(target, output_dir, args.sheet, args.convert_all)
    elif ext == ".csv":
        convert_csv_encoding(target, output_dir, args.encoding)
    else:
        print(f"ERROR: Unsupported file type: {ext}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
